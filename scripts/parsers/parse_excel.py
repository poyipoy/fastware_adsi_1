import os
import openpyxl
import json
import sys

base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../..'))
excel_path = os.path.join(base_dir, 'database', 'data', 'OK-ADASI_Mapping_JobPosition_Section_Dept.xlsx')
if not os.path.exists(excel_path):
    excel_path = 'OK-ADASI_Mapping_JobPosition_Section_Dept.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    sheet_name1 = 'Mapping (OK)'
    if sheet_name1 not in wb.sheetnames:
        matches = [s for s in wb.sheetnames if 'Mapping' in s]
        if matches:
            sheet_name1 = matches[0]
            print(f"Info: Using sheet '{sheet_name1}' instead of 'Mapping (OK)'")
        else:
            raise KeyError(f"Sheet 'Mapping (OK)' not found. Available sheets: {wb.sheetnames}")
            
    sheet1 = wb[sheet_name1]
    
    # Read specific cells
    notes = {
        'K7': sheet1['K7'].value,
        'J44': sheet1['J44'].value,
        'J74': sheet1['J74'].value,
        'L3': sheet1['L3'].value,
        'L5': sheet1['L5'].value
    }
    
    # Parse rows 3 to 95
    employees = []
    for row in range(3, 96):
        # We check column B (column 2) for fill color
        cell_b = sheet1.cell(row=row, column=2)
        fill_color = cell_b.fill.start_color.index if cell_b.fill and cell_b.fill.start_color else None
        
        emp = {
            'row': row,
            'no': sheet1.cell(row=row, column=1).value,
            'nama': sheet1.cell(row=row, column=2).value,
            'job_position': sheet1.cell(row=row, column=3).value,
            'section_job': sheet1.cell(row=row, column=4).value,
            'section_user': sheet1.cell(row=row, column=5).value,
            'section_head': sheet1.cell(row=row, column=6).value,
            'dept_job': sheet1.cell(row=row, column=7).value,
            'dept_head': sheet1.cell(row=row, column=8).value,
            'color': str(fill_color)
        }
        if emp['nama']: # Only add if name is not empty
            employees.append(emp)

    sheet_name2 = 'Update Section & Dept'
    if sheet_name2 not in wb.sheetnames:
        matches = [s for s in wb.sheetnames if 'Update' in s or 'Section' in s]
        if matches:
            sheet_name2 = matches[0]
            print(f"Info: Using sheet '{sheet_name2}' instead of 'Update Section & Dept'")
        else:
            raise KeyError(f"Sheet 'Update Section & Dept' not found. Available sheets: {wb.sheetnames}")
            
    sheet2 = wb[sheet_name2]
    official_sections = []
    for row in range(3, 18):
        sec = {
            'no': sheet2.cell(row=row, column=3).value,
            'section': sheet2.cell(row=row, column=4).value,
            'dept': sheet2.cell(row=row, column=5).value,
            'div': sheet2.cell(row=row, column=6).value,
        }
        if sec['section']:
            official_sections.append(sec)

    output = {
        'notes': notes,
        'employees': employees,
        'official_sections': official_sections
    }
    
    output_path = os.path.join(base_dir, 'database', 'data', 'parse_result.json')
    with open(output_path, 'w') as f:
        json.dump(output, f, indent=2)
    print("Success")
except Exception as e:
    print(f"Error: {e}")
